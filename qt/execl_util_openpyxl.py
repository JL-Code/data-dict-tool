#! /usr/bin/python
# -*- coding: utf-8 -*-

import logging

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

"""
    data 导出数据
    filepath 文件存放地址
    https://openpyxl.readthedocs.io/en/stable/pandas.html
"""


# TODO: sheet 列自动适应内容宽度 【1】
# TODO: sheet 序号文本居中 【3】
# TODO: 清理生成的空白 sheet 【2】
# TODO: 自定义表头样式 【3】
# TODO: 生成返回目录超链接 【3】
def build(data, filepath):
    logging.debug("目标路径: %s", filepath)
    try:
        wb = Workbook()
        catalog_name = "目录"
        index = append_row_to_sheet(wb, catalog_name, data[catalog_name], True, 0)
        for key in data:
            if key == "目录":
                continue
            index = append_row_to_sheet(wb, key, data[key], False, index + 1)
        """问题描述: https://github.com/xlwings/xlwings/issues/957"""
        """解决方案: https://github.com/xlwings/xlwings/pull/1372"""
        # wb.remove_sheet(wb.active)  # fix: 删除默认生成的空白 Sheet
        wb.save(filepath)

    except IOError as e:
        logging.debug("IOError:", e)


"""单元格增加超链接 
 https://stackoverflow.com/questions/39077661/adding-hyperlinks-in-some-cells-openpyxl
'=HYPERLINK("{}", "{}")'.format(link, "Link Name")
e.g. ws.cell(row=1, column=1).value = '=HYPERLINK("{}", "{}")'.format(link, "Link Name")
"""


def append_row_to_sheet(wb: Workbook, title, data, is_catalog: False, sheet_index: 0):
    """ 给目录添加行
    :param wb: Workbook 对象
    :param title:  title of the sheet
    :param data: 表格数据 Dict 类型
    :param is_catalog: 是否是目录 sheet，当是目录 sheet 是会给列字段增加超链接效果
    :param sheet_index: sheet 索引
    """
    column_widths = []
    logging.debug("创建 sheet %s", title)
    worksheet = wb.create_sheet(title=title, index=sheet_index)
    df = pd.DataFrame(data)
    rows = dataframe_to_rows(df, index=False, header=True)
    start_row_index = 0  # 0 是 header，数据行从 1 开始。

    if is_catalog:
        for row in rows:
            if start_row_index != 0:
                row[1] = '=HYPERLINK("#{}!A1", "{}")'.format(row[2], row[1])
            add_column_width(row, column_widths)
            worksheet.append(row)
            start_row_index += 1
    else:
        # 填充返回目录信息
        cl_df = pd.DataFrame([{"TEXT": '=HYPERLINK("#{}!A1", "{}")'.format('目录', '返回目录')}])
        cl_df_rows = dataframe_to_rows(cl_df, index=False, header=False)
        for row in cl_df_rows:
            worksheet.append(row)
        # 填充数据字典信息
        rows = dataframe_to_rows(df, index=False, header=True)
        for row in rows:
            add_column_width(row, column_widths)
            worksheet.append(row)

    # 手动设置列宽度
    for i, column_width in enumerate(column_widths):
        worksheet.column_dimensions[get_column_letter(i + 1)].width = column_width

    return wb.get_index(worksheet)


""" python openpyxl列宽自适应
column_widths = []
for row in data:
    for i, cell in enumerate(row):
        if len(column_widths) > i:
            if len(cell) > column_widths[i]:
                column_widths[i] = len(cell)
        else:
            column_widths += [len(cell)]

for i, column_width in enumerate(column_widths):
    worksheet.column_dimensions[get_column_letter(i+1)].width = column_width
https://leadscloud.github.io/313773/python-openpyxl%E5%88%97%E5%AE%BD%E8%87%AA%E9%80%82%E5%BA%94/
https://blog.csdn.net/crammy/article/details/120469646
"""


def add_column_width(row, column_widths):
    """ add_column_width
    :param row: sheet 数据行
    :param column_widths: 列宽度数组，存放每一列的宽度
    """
    try:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if obtain_len(cell) > column_widths[i]:
                    column_widths[i] = obtain_len(cell)
            else:
                column_widths += [obtain_len(cell)]
    except TypeError as e:
        logging.debug("IOError:", e)


def obtain_len(cell):
    return len(cell.encode('gbk')) + 2 if isinstance(cell, str) else len(str(cell)) + 2
